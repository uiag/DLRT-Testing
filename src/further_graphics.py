# -*- coding: utf-8 -*-
"""
Created on Tue Aug 29 14:25:49 2023

@author: jakob
"""
import openpyxl
import matplotlib.pyplot as plt
import numpy as np

from matplotlib import cm
from matplotlib.ticker import LinearLocator

plt.rcParams["figure.figsize"] = (10,14.14)
plt.rcParams.update({'font.size': 10})

file = openpyxl.load_workbook('data.xlsx')
sheet = file["dataAll"]

i = 2
allEntries = []
while sheet.cell(row = i, column = 1).value != "EOI":
    entry = [sheet.cell(row = i, column = 1).value, sheet.cell(row = i, column = 2).value, sheet.cell(row = i, column = 3).value, float(sheet.cell(row = i, column = 4).value)]
    allEntries.append(entry)
    i += 1
    
i = 0
while allEntries[i][1] == "Drawing":
    plt.plot(allEntries[i][2], allEntries[i][3], 'yo')
    i += 1
while allEntries[i][1] == "Games":
    plt.plot(allEntries[i][2], allEntries[i][3], 'go')
    i += 1
while allEntries[i][1] == "Language":
    plt.plot(allEntries[i][2], allEntries[i][3], 'bo')
    i += 1
while allEntries[i][1] == "Vision":
    plt.plot(allEntries[i][2], allEntries[i][3], 'ro')
    i += 1
while i < len(allEntries):
    plt.plot(allEntries[i][2], allEntries[i][3], 'mo')
    i += 1

plt.yscale("log")    
plt.show()

plt.rcParams["figure.figsize"] = (16,9)
plt.rcParams.update({'font.size': 20})

handwriting = [[1998, 1998, 2002, 2003, 2006, 2010, 2012, 2013, 2018], [-100, -80, -49, -27, -25, -20, -5, -2, 3]]
speech = [[1998, 2011, 2013, 2014, 2015, 2015, 2016, 2016, 2017, 2018], [-100, -67, -51, -28, -27, -9, -5, -1, 1, 2]]
image = [[2009, 2012, 2014, 2014, 2015, 2016, 2018, 2019, 2020],[-100, -44, -9, -7, 1, 7, 12, 10, 18]]
reading = [[2016, 2016, 2017, 2017, 2018, 2019, 2020], [-100, -32, -29, -9, 6, 19, 19]]
language = [[2018, 2018, 2019, 2019, 2019, 2019, 2020, 2020],[-100, -68, -64, -25, 0, 4, 8, 12]] 
human = [[2006, 2020], [0, 0]]

plt.plot(handwriting[0], handwriting[1], 'r-', label="Handwriting recognition")
plt.plot(speech[0], speech[1], 'g-', label="Speech recognition")
plt.plot(image[0], image[1], 'b-', label="Image recognition")
plt.plot(reading[0], reading[1], 'y-', label="Reading comprehension")
plt.plot(language[0], language[1], 'm-', label="Language understanding")
plt.plot(human[0], human[1], 'k-')
plt.legend()
plt.show()


ax = plt.subplot(projection="3d", computed_zorder=False)

step = 0.12
x = -1
y = -0.1
data = [[-1,-0.1,2.2937]]
for i in range(8):
    xNew = x - step * (8*x-8.4*x**3+2*x**5+y)
    yNew = y - step * (x-8*y+16*y**3)
    x = xNew
    y = yNew
    z = ((4-2.1*x**2+(x**4/3))*x**2+x*y+(-4+4*y**2)*y**2)
    data.append([xNew, yNew, z])
    x = xNew
    y = yNew
    
step = 0.001
x = -1
y = -0.1
data2 = [[-1,-0.1,2.2937]]
for i in range(1000):
    xNew = x - step * (8*x-8.4*x**3+2*x**5+y)
    yNew = y - step * (x-8*y+16*y**3)
    x = xNew
    y = yNew
    z = ((4-2.1*x**2+(x**4/3))*x**2+x*y+(-4+4*y**2)*y**2)
    data2.append([xNew, yNew, z])
    x = xNew
    y = yNew
    
# Make data.
X = np.arange(-2, 2, 0.1)
Y = np.arange(-1, 1, 0.1)
X, Y = np.meshgrid(X, Y)
Z = ((4-2.1*X**2+(X**4/3))*X**2+X*Y+(-4+4*Y**2)*Y**2)

# Plot the surface.
surf = ax.plot_surface(X, Y, Z, cmap="viridis",
                       linewidth=0, zorder=-1)

ax.scatter(data[0][0], data[0][1], data[0][2], color="red", zorder=1, label="gradient descent")
for el in data[1:]:
    ax.scatter(el[0], el[1], el[2], color="red", zorder=1)

ax.scatter(data2[0][0], data2[0][1], data2[0][2], color="black", zorder=1, label="gradient flow")
for el in data2[1:]:
    ax.scatter(el[0], el[1], el[2], color="black", s=10, zorder=0)


# Customize the z axis.
ax.set_zlim(-2, 6)
#ax.zaxis.set_major_locator(LinearLocator(10))
# A StrMethodFormatter is used automatically
#ax.zaxis.set_major_formatter('{x:.02f}')

# Add a color bar which maps values to colors.
plt.legend()
plt.show()